from fastapi import APIRouter
from fastapi import UploadFile
from fastapi import File
from io import BytesIO
from database.mongodb import db
from datetime import datetime, timezone

from database.mongodb import db

from openpyxl import load_workbook

router = APIRouter()


@router.post("/hostel/add-students")

async def add_students(
    file: UploadFile = File(...)
):

    contents = await file.read()

    workbook = load_workbook(
    BytesIO(contents)
)

    sheet = workbook.active

    success_count = 0

    duplicate_count = 0

    error_count = 0

    errors = []

    branch_map = {

        "BCS": "CSE",
        "BME": "ME",
        "BEC": "ECE",
        "BSM": "SM",
        "BDS": "Design"

    }

    for row in range(
        2,
        sheet.max_row + 1
    ):

        email = str(
            sheet[f"A{row}"].value or ""
        ).strip()

        name = str(
            sheet[f"B{row}"].value or ""
        ).strip()

        room = str(
            sheet[f"C{row}"].value or ""
        ).strip()

        if (
            email == "" and
            name == "" and
            room == ""
        ):
            continue

        issues = []

        if "@iiitdmj.ac.in" not in email:

            issues.append(
                "Invalid Email"
            )

        roll = email.split("@")[0].upper()

        if len(roll) < 5:

            issues.append(
                "Invalid Roll"
            )

        branch = "UNKNOWN"

        if len(roll) >= 5:

            code = roll[2:5]

            if code in branch_map:

                branch = branch_map[code]

            else:

                issues.append(
                    "Unknown Branch"
                )

        existing = db.students_excel.find_one({

            "roll_no": roll

        })

        if existing:

            duplicate_count += 1

            continue

        if issues:

            error_count += 1

            errors.append({

                "row": row,

                "roll": roll,

                "issue":
                    ", ".join(issues)

            })

            continue

        db.students_excel.insert_one({

            "name": name,

            "roll_no": roll,

            "room": room,

            "branch": branch,

            "hostel": "Hostel 1",

            "email": email,

            "created_at": datetime.now().strftime("%H:%M:%S %d-%m-%Y")

        })

        success_count += 1

    return {

        "success_count":
            success_count,

        "duplicate_count":
            duplicate_count,

        "error_count":
            error_count,

        "errors":
            errors

    }